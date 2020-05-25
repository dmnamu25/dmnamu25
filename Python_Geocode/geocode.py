import sys
from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import json
import os.path

#윈도우 맥 리눅스등 현재 디렉토리에서 파일 찾을려고
folder = os.getcwd()
filename = '통합 문서1.xlsx'
fullname = os.path.join(folder,filename)



#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook(fullname, data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet']

# 엑셀에 쓰기 
write_wb = Workbook()
#이름이 있는 시트를 생성
#write_ws = write_wb.create_sheet('지오코드')
 
#Sheet1에다 입력
write_ws = write_wb.active
 
print('\n-----변환을 시작합니다.-----')
all_values = []
row_cnt =0
success_cnt=0
fail_cnt=0
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
    #print(row_value)
    address=" ".join(row_value)
    r =requests.get('http://apis.vworld.kr/new2coord.do?q='+address+'&apiKey=767B7ADF-10BA-3D86-AB7E-02816B5B92E9&domain=http://map.vworld.kr/&output=json')
    #print(r.json())
    data = r.json()
    #셀 단위로 추가
    row_cnt = row_cnt+1
    #주소
    write_ws.cell(row_cnt,1,address)
    try:
    #위도
        latitude=data['EPSG_4326_Y']
        write_ws.cell(row_cnt,2,latitude)
    #경도
        longtitude=data['EPSG_4326_X']
        write_ws.cell(row_cnt,3,longtitude)
        success_cnt+=1
    except Exception: 
        write_ws.cell(row_cnt,2,'위경도 가져오기 실패/ 주소를 확인하세요')
        fail_cnt+=1
    finally:
        print(str(row_cnt))

#나머지 파일저장
print(str(success_cnt)+"건 변환성공// "+str(fail_cnt)+"건 변환실패")
write_wb.save(fullname)



